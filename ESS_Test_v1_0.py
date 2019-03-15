import visa
import time
from datetime import time as tm
import math
from openpyxl import Workbook
from openpyxl import load_workbook
#from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    Paragraph,
    ParagraphProperties,
    CharacterProperties,
    RichTextProperties,
    Font,
    RegularTextRun,
    )
"""
# start of Untitled

rm = visa.ResourceManager()

#Connects to Keithley 2110 Digit Multimeter VIA USB Type B
MODEL_2110 = rm.open_resource('USB0::0x05E6::0x2110:s:1374051::0::INSTR')

# Clears and Resets the Multimeter
idn = MODEL_2110.query('*IDN?')
MODEL_2110.write('*RST')

#Set Voltage and Thermocouple Measurements
MODEL_2110.write(':SENSe:VOLTage:DC:RANGe:AUTO %s' % ('ON'))
MODEL_2110.write('FUNCtion2 "TCOuple"')

#Read Thermocouple Measurements
temp_values = MODEL_2110.query_ascii_values(':READ?')
readings = temp_values[0:]


MODEL_2110.close()
rm.close()

# end of Untitled
"""

class TEST_XM:
    def __init__(self):
        self.rm = visa.ResourceManager()
        self.ports = None
        self.models = None
        self.start_time = None
        self.elapsed_time = None

    def get_rm(self):
        return self.rm

    def set_ports(self):
        self.ports = self.rm.list_resources()
        #mixer.init()
        #mixer.music.load('C:\Users\samscheer\Desktop\TEST\TENNEY\sound.mp3')
        print(self.ports)
        #self.models = [self.rm.open_resource(port) for port in self.ports]
        try:
            self.models = self.rm.open_resource('USB0::0x05E6::0x2110::1374051::INSTR')
        except:
            print('error: usb?')
            self.models = self.rm.open_resource(self.ports[0])
        return self.models

    def get_ports(self):
        return self.ports
        #return self.models

    """
    def reset(self, model):
        model.write('*RST')

    def set_mode(self, model):
        model.write(':SENSe:VOLTage:DC:RANGe:AUTO %s' % ('ON'))
        model.write('FUNCtion2 "TCOuple"')

    def measure(self, model):
        temp_values = model.query_ascii_values(':READ?')
        readings = temp_values[0:]
        return readings
    """

    def reset(self):
        self.models.write('*RST')

    def set_mode(self):
        self.models.write(':SENSe:VOLTage:DC:RANGe:AUTO %s' % ('ON'))
        self.models.write('FUNCtion2 "TCOuple"')

    def measure(self):
        temp_values = self.models.query_ascii_values(':READ?')
        readings = temp_values[0:]
        #if readings[0] <= 0:
            #mixer.music.play()
            #print('fail')
            #raise SystemExit #######################33
        return readings

    def close(self):
        self.models.close()
        self.rm.close()

    def start_timer(self):
        self.start_time = time.time()

    def get_time(self):
        current_time = time.time()- self.start_time
        formatted_time = self.format_time(current_time)
        #return current_time
        sec = math.floor(current_time)
        time_string = self.string_time(current_time)
        time_list = [formatted_time, sec, time_string]
        return time_list

    def end_time(self):
        self.elapsed_time = time.time() - self.start_time
        formatted_time = self.format_time(math.floor(self.elapsed_time))
        #return self.elapsed_time
        #formatted_time = math.floor(self.elapsed_time)
        return formatted_time

    def format_time(self, input_time):
        formatted_time = tm(math.floor(input_time)//3600, (math.floor(input_time)%3600)//60, math.floor(input_time)%60) # h, m, s
        #formatted_time = time.strftime("%H:%M:%S", time.gmtime(input_time))
        return formatted_time

    def string_time(self, input_time):
        formatted_time = time.strftime("%H:%M:%S", time.gmtime(input_time))
        return formatted_time
"""    
    def start_timer(self):
        self.start_time = datetime.datetime.now()

    def get_time(self):
        current_time = datetime.datetime.now().time() - self.start_time
        formatted_time = current_time.time()
        return formatted_time

    def end_time(self):
        self.elapsed_time = datetime.datetime.now().time() - self.start_time
        formatted_time = self.elapsed_time.time()
        return formatted_time

      
"""

class TEST_XL():
    def __init__(self, fileName, sheetName):
        self.wb = None
        self.ws = None
        self.chart = None
        self.aux_chart = None
        self.fileName = fileName
        self.sheetName = sheetName

    def set_wb(self):
        create_flag = False
        try:
            self.wb = load_workbook(self.fileName)
            self.wb.save(self.fileName)
            load_flag = True
        except:
            load_flag = False
            #print('Worksheet loading failed...')
            
        if load_flag is False:
            try:
                self.wb = Workbook()
                self.wb.save(self.fileName)
                create_flag = True
            except:
                create_flag = False
                #print('Worksheet creation failed...')

        if create_flag is False and load_flag is False:
            print("ERROR: Is " + self.fileName + " already opened?\n")
            raise SystemExit
        else:
            if create_flag is True:
                print(self.fileName + ' created successfully!\n')
            elif load_flag is True:
                print(self.fileName + ' loaded successfully!\n')

    def get_wb(self):
        return self.wb

    def set_ws(self):
        self.ws = self.wb.active
        self.ws.title = self.sheetName

    def get_ws(self):
        return self.ws

    def set_file(self, fileName):
        self.fileName = fileName

    def get_file(self):
        return self.fileName

    def set_sheet(self, sheetName):
        self.sheetName = sheetName

    def get_sheet(self):
        return self.sheetName

    def format_print_margins(self):
        self.ws.page_setup.orientation = self.ws.ORIENTATION_LANDSCAPE
        self.ws.page_setup.scale = 80
        self.ws.page_margins.left = .25
        self.ws.page_margins.right = .25
        self.ws.page_margins.top = .75
        self.ws.page_margins.bottom = .75
        self.ws.page_margins.header = .3
        self.ws.page_margins.footer = .3

    def page_header(self):
        self.ws.column_dimensions["A"].width = 12
        self.ws.column_dimensions["B"].width = 12
        self.ws.column_dimensions["C"].width = 12
        self.ws.column_dimensions["D"].width = 12

        header = ['TIME(s)', 'VOLTAGE', 'TEMP(C)']

        self.write_xl(header)

    def write_xl(self, data):
        #print(data)
        self.ws.append(data)
        self.save_xl()

    def save_xl(self):
        self.wb.save(self.fileName)

    def create_graph(self):
        self.chart = ScatterChart()
        self.chart.title = "ESS TEST"
        self.chart.style = 13
        self.chart.x_axis.title = 'Elapsed Time'
        self.chart.x_axis.number_format = 'h:mm:ss'
        self.chart.y_axis.title = 'RF Detector Voltage (V)'
        self.chart.height = 10
        self.chart.width = 15
        #self.chart.legend = True

        self.chart.x_axis.scaling.min = 0
        #self.chart.x_axis.scaling.max = self.ws['A' + str(self.get_max_row())].value
        #print(self.ws['A' + str(self.get_max_row())].value)

        #self.chart.y_axis.scaling.min = 0
        #self.chart.y_axis.scaling.max = 0.2
        
        self.aux_chart = ScatterChart()
        self.aux_chart.y_axis.axId = 200
        self.aux_chart.y_axis.title = 'Temp (C)'
        self.aux_chart.y_axis.majorGridlines = None
        
        self.format_chart(self.chart)
        self.format_chart(self.aux_chart)
        
    def format_chart(self, target_chart):
        font = Font(typeface = "Calibri")
        cp8 = CharacterProperties(latin = font, sz = 800, b = False)
        cp10 = CharacterProperties(latin = font, sz = 1000, b = False)
        cp12 = CharacterProperties(latin = font, sz = 1200, b = False)

        pp8 = ParagraphProperties(defRPr = cp8)
        pp10 = ParagraphProperties(defRPr = cp10) # title font
        pp12 = ParagraphProperties(defRPr = cp12)

        rtp8 = RichText(p = [Paragraph( pPr = pp8, endParaRPr = cp8)]) # axis font
        rtp10 = RichText(p = [Paragraph( pPr = pp10, endParaRPr = cp10)]) 
        rtp12 = RichText(p = [Paragraph( pPr = pp12, endParaRPr = cp12)])

        target_chart.x_axis.txPr = rtp8
        target_chart.y_axis.txPr = rtp8
        self.chart.x_axis.title.tx.rich.p[0].pPr = pp10
        target_chart.y_axis.title.tx.rich.p[0].pPr = pp10
        self.chart.title.tx.rich.p[0].pPr = pp12

    def select_data(self):
        last_row = self.get_max_row()
        xvalues = Reference(self.ws, min_col = 1, min_row = 2, max_row = last_row)
        for i in range (2, 4):  ###CHANGE THESE VALUES
            yvalues = Reference(self.ws, min_col = i, min_row = 1, max_row = last_row)
            series = Series(yvalues, xvalues, title_from_data = True)
            if (i == 2):
                self.chart.series.append(series)
                self.chart.series[0].graphicalProperties.line.solidFill = 'FF0000'
            else :
                self.aux_chart.series.append(series)
                self.aux_chart.series[0].graphicalProperties.line.solidFill = '00AAAA'

        self.aux_chart.y_axis.crosses = 'max'
        self.chart += self.aux_chart

        self.ws.add_chart(self.chart, get_column_letter(self.get_max_col() + 2) + str(1))

        self.save_xl()
        
    def get_max_row(self):
        max_row = self.ws.max_row
        return max_row

    def get_max_col(self):
        max_col = self.ws.max_column
        return max_col
    
def main():
    xm = TEST_XM()
    visa_coms = xm.set_ports()
#    xm.set_mode(visa_coms)
    xm.set_mode()

    xl = TEST_XL('pldro8125.xlsx', 'Sheet 1')
    xl.set_wb()
    xl.set_ws()

    #xl.format_print_margins()
    #xl.page_header()
    
    xm.start_timer()

    #hour = 1
    #mins = 60
    #secs = 60
    
    step_2 = 15
    step_3 = 45
    step_4 = 25
    step_5 = 45
    step_6 = 15

    step_add = 20

    time_min = 20
    time_sec = time_min * 60

    cycle = (step_2 + step_3 + step_4 + step_5+step_6) * 60
    print(cycle)
    cycle_3 = (cycle * 3) + (25*60)
    #time_sec = 1 * hour * mins * secs

    #for x in range (0, time_sec + 30, 30):
    while True:
        values = xm.measure()
        voltage = abs(values[0]) 
        if (voltage <= 0):
            print("fail")
        temp = values[1]
        time_elapsed = xm.get_time()    
        if (time_elapsed[1] > cycle_3):
            break
        data = [time_elapsed[0], float(voltage), float(temp)]
        print([time_elapsed[2], float(voltage), float(temp)])
        xl.write_xl(data)
        time.sleep(30)

    xm.end_time()
    
    xm.close()

    xl.create_graph()
    print('Writing to file...\n')
    time.sleep(5)
    xl.select_data()
    xl.save_xl()
    print('File saved successfully!\n')
        
main()

