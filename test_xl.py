from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import (
    Paragraph,
    ParagraphProperties,
    CharacterProperties,
    RichTextProperties,
    Font,
    RegularTextRun,
    )

class TEST_XL():
    def __init__(self,fileName):
        self.wb = None
        self.ws = None
        self.chart = None
        self.chart.title = None
        self.chart.y_axis.title = None
        self.aux_chart = None
        self.sheet_name = 'ESS Test'
        self.fileName = fileName

    def set_wb(self):
        create_flag = False
        #LOAD XL
        try:
            self.wb = load_workbook(self.fileName)
            self.wb.save(self.fileName)
            load_flag = True
        except:
            load_flag = False
            #print('Worksheet loading failed...')

        #CREATE XL
        if load_flag is False:
            try:
                self.wb = Workbook()
                create_flag = True
            except:
                create_flag = False
                #print('Worksheet creation failed...')

        #XL ALREADY OPENED
        if create_flag is False and load_flag is False:
            print("ERROR: Is " + self.fileName + " already opened?\n")
            raise SystemExit
        else:
            if create_flag is True:
                self.ws = self.wb.active
                self.ws.title = self.sheet_name
                print(self.fileName + ' created successfully!\n')
            elif load_flag is True:
                self.ws = self.wb.active
                print(self.fileName + ' loaded successfully!\n')
        
            self.wb.save(self.fileName)            

    def format_print_margins(self):
        self.ws.page_setup.orientation = self.ws.ORIENTATION_LANDSCAPE
        self.ws.page_setup.scale = 80
        self.ws.page_margins.left = .25
        self.ws.page_margins.right = .25
        self.ws.page_margins.top = .75
        self.ws.page_margins.bottom = .75
        self.ws.page_margins.header = .3
        self.ws.page_margins.footer = .3

    def page_header(self):
        self.ws.column_dimensions["A"].width = 12
        self.ws.column_dimensions["B"].width = 12
        self.ws.column_dimensions["C"].width = 12

        header = ['Time', 'Voltage', 'Temp.(C)']

        self.write_xl(header)

    def write_xl(self, data):
        self.ws.append(data)

        self.save_xl()

    def save_xl(self):
        self.wb.save(self.fileName)

    def create_graph(self):
        self.chart = ScatterChart()
        self.chart.style = 13
        self.chart.height = 10
        self.chart.width = 15

        self.chart.x_axis.title = 'Elapsed Time(hours)'
        self.chart.x_axis.number_format = 'h:mm:ss'
        self.chart.x_axis.scaling.min = 0

        self.aux_chart = ScatterChart()
        self.aux_chart.y_axis.axId = 200
        self.aux_chart.y_axis.title = 'Temp (C)'
        self.aux_chart.y_axis.majorGridlines = None
        
        self.format_chart(self.chart)
        self.format_chart(self.aux_chart)
        
    def format_chart(self, target_chart):
        font = Font(typeface = "Calibri")
        cp8 = CharacterProperties(latin = font, sz = 800, b = False)
        cp10 = CharacterProperties(latin = font, sz = 1000, b = False)
        cp12 = CharacterProperties(latin = font, sz = 1200, b = False)

        pp8 = ParagraphProperties(defRPr = cp8)
        pp10 = ParagraphProperties(defRPr = cp10) # title font
        pp12 = ParagraphProperties(defRPr = cp12)

        rtp8 = RichText(p = [Paragraph( pPr = pp8, endParaRPr = cp8)]) # axis font
        rtp10 = RichText(p = [Paragraph( pPr = pp10, endParaRPr = cp10)]) 
        rtp12 = RichText(p = [Paragraph( pPr = pp12, endParaRPr = cp12)])

        target_chart.x_axis.txPr = rtp8
        target_chart.y_axis.txPr = rtp8
        target_chart.x_axis.title.tx.rich.p[0].pPr = pp10
        target_chart.y_axis.title.tx.rich.p[0].pPr = pp10
        target_chart.title.tx.rich.p[0].pPr = pp12

    def graph_data(self):
        last_row = self.get_max_row()
        xvalues = Reference(self.ws, min_col = 1, min_row = 2, max_row = last_row)
        for i in range (2, 4):  ###CHANGE THESE VALUES
            yvalues = Reference(self.ws, min_col = i, min_row = 1, max_row = last_row)
            series = Series(yvalues, xvalues, title_from_data = True)
            if (i == 2):
                self.chart.series.append(series)
                self.chart.series[0].graphicalProperties.line.solidFill = 'FF0000'
            else :
                self.aux_chart.series.append(series)
                self.aux_chart.series[0].graphicalProperties.line.solidFill = '00AAAA'

        self.aux_chart.y_axis.crosses = 'max'
        self.chart += self.aux_chart

        self.ws.add_chart(self.chart, get_column_letter(self.get_max_col() + 2) + str(1))

        self.save_xl()
        
    def get_max_row(self):
        max_row = self.ws.max_row
        return max_row

    def get_max_col(self):
        max_col = self.ws.max_column
        return max_col
